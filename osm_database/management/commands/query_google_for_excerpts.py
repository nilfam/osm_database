import os
import pathlib
import pickle
import sys
import urllib

import pandas as pd
from bs4 import BeautifulSoup
from django.core.management import BaseCommand
from openpyxl import load_workbook
from progress.bar import Bar
from selenium import webdriver
from sortedcontainers import SortedDict

from osm_database.management.util.browser_wrapper import BrowserWrapper
from root.utils import send_capcha_unsolve_limit_reached


current_dir = os.path.dirname(os.path.abspath(__file__))
script_name = os.path.split(__file__)[1][0:-3]
dir_parts = current_dir.split('/')
cache_dir = os.path.join('/'.join(dir_parts[0:dir_parts.index('management')]), 'cache', script_name)


def get_result_stats(body):
    result_stats = body.select('#result-stats')[0]
    no_results_indicator = body.select('.HlosJb.bOkdDe')
    if len(no_results_indicator) > 0:
        return 0
    nobrs = result_stats.select('nobr')
    for nobr in nobrs:
        nobr.decompose()
    result_stats_text = result_stats.text

    try:
        return int(''.join([s for s in result_stats_text if s.isdigit()]))
    except:
        print('Unrecognised expression: {}'.format(result_stats_text))
        return 'Not Found'


def get_next_pages_links(body):
    pages = []
    nav = body.select('div[role="navigation"]')[-1]
    page_links = nav.select('table td a')
    for a in page_links:
        label = a.get('aria-label')
        if label is not None:
            url = 'https://google.com' + a.get('href')
            pages.append((label, url))
    return pages


def get_searched_items(body, items):
    gs = body.select('#search div.g')
    for g in gs:
        link = g.select('.yuRUbf a')[0].get('href')
        excerpt = g.select('.aCOpRe')[0].text
        items.append(Item(link, excerpt))


class CookiesPopulationRequiredException(Exception):
    def __init__(self, cookies):
        super(CookiesPopulationRequiredException, self).__init__()
        self.cookies = cookies


class Item:
    def __init__(self, url, excerpt):
        self.url = url
        self.excerpt = excerpt


class Page:
    def __init__(self, query, url, name):
        self.items = []
        self.url = url
        self.name = name
        self.query = query
        self.result_count = 0
        self.finalised = False
        self.cache_dir = os.path.join(cache_dir, 'html', self.query.replace(' ', '_').replace('/', '_or_'))
        pathlib.Path(self.cache_dir).mkdir(parents=True, exist_ok=True)

    def parse_html_get_next_pages(self):
        cache_file_path = os.path.join(self.cache_dir, '{}.html'.format(self.name.replace(' ', '_')))

        if not os.path.isfile(cache_file_path):
            raise Exception('File {} doesn\'t exist. Please query it first')

        with open(cache_file_path, 'r') as f:
            response = f.read()

        soup = BeautifulSoup(response)
        body = soup.find('body')
        try:
            self.result_count = get_result_stats(body)
        except:
            self.result_count = -1
        finally:
            if self.result_count > 0:
                get_searched_items(body, self.items)
                next_pages = get_next_pages_links(body)
            else:
                next_pages = []

        return next_pages

    def make_query_get_next_pages(self, browser_wrapper):
        cache_file_path = os.path.join(self.cache_dir, '{}.html'.format(self.name.replace(' ', '_')))

        if os.path.isfile(cache_file_path):
            with open(cache_file_path, 'r') as f:
                response = f.read()

        else:
            response = browser_wrapper.make_query_retrial_if_fail(self.url)
            with open(cache_file_path, 'w', encoding='utf-8') as htmlfile:
                htmlfile.write(response)

        soup = BeautifulSoup(response)
        body = soup.find('body')
        try:
            self.result_count = get_result_stats(body)
        except:
            self.result_count = -1
        finally:
            if self.result_count > 0:
                # get_searched_items(body, self.items)
                try:
                    next_pages = get_next_pages_links(body)
                except:
                    next_pages = []
            else:
                next_pages = []

        return next_pages


class QueryResult:
    def __init__(self, query):
        self.query = query
        self.cache_dir = os.path.join('data', self.query.replace(' ', '_'))
        self.cache_file = os.path.join(self.cache_dir, 'query_result-2.pkl')
        pathlib.Path(self.cache_dir).mkdir(parents=True, exist_ok=True)

        if os.path.isfile(self.cache_file):
            with open(self.cache_file, 'rb') as f:
                _page = pickle.load(f)
                self.pages = _page.pages
                self.not_queried_page_names = _page.not_queried_page_names
                self.num_pages_queries = getattr(_page, 'num_pages_queries', 0)
        else:
            self.pages = {}
            self.not_queried_page_names = []
            self.num_pages_queries = 0
            query_plus = urllib.parse.quote_plus(query)
            first_page_url = "https://google.com/search?q=\"{}\"&hl=en&num=100".format(query_plus)
            first_page = Page(query, first_page_url, 'Page 1')
            self.not_queried_page_names.append('Page 1')
            self.pages['Page 1'] = first_page

    def make_query_recursive(self, max_pages, browser_wrapper):
        while len(self.not_queried_page_names) > 0 and self.num_pages_queries < max_pages:
            page_name = self.not_queried_page_names[0]
            page_to_query = self.pages[page_name]
            next_pages = page_to_query.make_query_get_next_pages(browser_wrapper)

            for next_page_name, next_page_url in next_pages:
                if next_page_name not in self.pages:
                    self.pages[next_page_name] = Page(self.query, next_page_url, next_page_name)
                    self.not_queried_page_names.append(next_page_name)

            self.not_queried_page_names = self.not_queried_page_names[1:]
            self.num_pages_queries += 1

            with open(self.cache_file, 'wb') as f:
                pickle.dump(self, f)

    def parse_html_recursive(self, max_pages):
        while len(self.not_queried_page_names) > 0 and self.num_pages_queries < max_pages:
            page_name = self.not_queried_page_names[0]
            page_to_query = self.pages[page_name]
            next_pages = page_to_query.parse_html_get_next_pages()

            for next_page_name, next_page_url in next_pages:
                if next_page_name not in self.pages:
                    self.pages[next_page_name] = Page(self.query, next_page_url, next_page_name)
                    self.not_queried_page_names.append(next_page_name)

            self.not_queried_page_names = self.not_queried_page_names[1:]
            self.num_pages_queries += 1

            with open(self.cache_file, 'wb') as f:
                pickle.dump(self, f)

    def export_excel(self):
        export_dir = 'export'
        export_file = os.path.join(export_dir, '{}.xlsx'.format(self.query.replace(' ', '_')))
        pathlib.Path(export_dir).mkdir(parents=True, exist_ok=True)

        df = pd.DataFrame(columns=['Page No', 'Excerpt', 'Link'])
        ind = 0
        for page_name, page in self.pages.items():
            for item in page.items:
                df.loc[ind] = [page_name, item.excerpt, item.url]
                ind += 1

        with pd.ExcelWriter(export_file, mode='w') as writer:
            df.to_excel(writer, startrow=0)

        book = load_workbook(export_file)
        ws = book.active
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

        book.save(export_file)


class Command(BaseCommand):
    def __init__(self):
        super().__init__()
        self.server = None
        self.context = None
        self.cache_dir = cache_dir
        pathlib.Path(cache_dir).mkdir(parents=True, exist_ok=True)

        self.query_result_cache_file = os.path.join(self.cache_dir, 'query_result-2.pkl')
        self.query_result_cache_file_bak = os.path.join(self.cache_dir, 'query_result-2.pkl.bak')

        self.expressions = SortedDict()

        if os.path.isfile(self.query_result_cache_file):
            with open(self.query_result_cache_file, 'rb') as f:
                try:
                    _query = pickle.load(f)
                    self.successful_queried_expressions = _query['successful_queried_expressions']
                    self.pages = _query['pages']
                except:
                    self.successful_queried_expressions = set()
                    self.pages = []
        else:
            self.successful_queried_expressions = set()
            self.pages = []

        self.browser_wrapper = BrowserWrapper(cache_dir)

    def add_arguments(self, parser):
        parser.add_argument('--file', action='store', dest='file', required=True, type=str)
        parser.add_argument('--auto', action='store_true', dest='automode', default=False)

    def populate_expressions_from_excel(self, file):
        xl = pd.ExcelFile(file)

        for sheet_name in xl.sheet_names:
            if sheet_name.startswith('#'):
                continue
            df = xl.parse(sheet_name, keep_default_na=False)
            df = df.fillna('')

            prepositions = []
            locatums = []

            for row_num, row in df.iterrows():
                preposition = row['Preposition']
                locatum = row['Locatum']
                if preposition:
                    prepositions.append(preposition)
                if locatum:
                    locatums.append(locatum)

            for preposition in prepositions:
                for locatum in locatums:
                    expression = locatum + ' ' + preposition + ' ' + sheet_name
                    self.expressions[expression] = (locatum, preposition, sheet_name)

    def make_query(self):
        expression_index = 0
        expressions = list(self.expressions)[15000:]
        expressions_count = len(expressions)
        for expression in expressions:
            print("Querying expression #{}/{}".format(expression_index, expressions_count))
            query_result = QueryResult(expression)
            query_finished = False
            while not query_finished:
                try:
                    query_result.make_query_recursive(max_pages=100, browser_wrapper=self.browser_wrapper)
                    query_finished = True
                except Exception as e:
                    self.browser_wrapper.reload()
            expression_index += 1

    def save(self):
        with open(self.query_result_cache_file_bak, 'wb') as f:
            cache = {'successful_queried_expressions': self.successful_queried_expressions, 'pages': self.pages}
            pickle.dump(cache, f)

        os.rename(self.query_result_cache_file_bak, self.query_result_cache_file)

    def get_name_to_counts(self):
        name_to_counts = {}
        for page in self.pages:
            count = page.extract_count()
            name_to_counts[page.query] = count
        return name_to_counts

    def parse_html(self):
        queries = {}
        for expression, (locatum, preposition, sheet_name) in self.expressions.items():
            query_result = QueryResult(expression)
            query_result.parse_html_recursive(max_pages=100)
            queries[expression] = (locatum, preposition, sheet_name, query_result)
        return queries

    def produce_report(self, queries):
        headings = ['Locatum', 'Preposition', 'Relatum', 'Page #', 'URL', 'Num results', 'Item Link', 'Item Excerpt']
        df = pd.DataFrame(columns=headings)
        index = 0
        bar = Bar('Exporting Excel file', max=len(queries))
        for expression, (locatum, preposition, sheet_name, query_result) in queries.items():
            total_page_count = len(query_result.pages)
            page_index = 1
            for page_name, page in query_result.pages.items():
                url = page.url
                page_number = '{}/{}'.format(page_index, total_page_count)
                num_results = page.result_count
                page_index += 1
                if len(page.items) == 0:
                    row = [locatum, preposition, sheet_name, page_number, url, num_results, '', '']
                    df.loc[index] = row
                    index += 1
                else:
                    for item in page.items:
                        row = [locatum, preposition, sheet_name, page_number, url, num_results, item.url, item.excerpt]
                        df.loc[index] = row
                        index += 1
            bar.next()
        bar.finish()

        export_file_path = os.path.join(self.cache_dir, 'export.xlsx')
        with pd.ExcelWriter(export_file_path, mode='w') as writer:
            df.to_excel(writer)

        book = load_workbook(export_file_path)
        ws = book.active
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

        book.save(export_file_path)

    def handle(self, *args, **options):
        file = options['file']
        self.browser_wrapper.auto_solve_captcha = options['automode']
        self.populate_expressions_from_excel(file)

        self.make_query()

        # queries = self.parse_html()
        # self.produce_report(queries)

        # self.test_2captcha()

        x = 0
